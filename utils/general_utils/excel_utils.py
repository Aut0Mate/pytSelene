from openpyxl import load_workbook


def get_all_rows_from_excel(file: str, work_sheet: str):
    """
    Gets all the rows, including the 1st Header row, from an excel worksheet
    :param file:
    :param work_sheet:
    :return: list of tuples containing cell values
    """
    wb = load_workbook(filename=file)
    ws = wb[work_sheet]
    all_rows = []
    for row in range(1, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row, col).value)
        all_rows.append(tuple(row_data))

    return all_rows


def get_data_from_excel_for_parameterize(file: str, work_sheet: str):
    """
    Returns the data from excel as a tuple of size 2 containing a tuple of headers and List of tuples(rows fom excel)
    :param file:
    :param work_sheet:
    :return: Tuple[Tuple, List[ANY]]
    """
    all_rows = get_all_rows_from_excel(file=file, work_sheet=work_sheet)
    headers = ",".join(all_rows[0])
    data = all_rows[1:]
    return headers, data
